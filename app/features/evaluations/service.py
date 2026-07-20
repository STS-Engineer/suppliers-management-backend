"""Batch evaluation ingestion and scheduling service."""

from __future__ import annotations

import calendar
import csv
import io
from datetime import date, datetime, timedelta
from typing import Any, Dict, List, Optional

from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.exceptions import AppException
from app.db.models import (
    AvocarbonSite,
    Classification,
    EvaluationCycle,
    ScoreCard,
    SupplierDevelopmentPlan,
    SupplierSiteRelation,
    SupplierStatusHistory,
    SupplierUnit,
)
from app.shared.utils.email.email_service import send_email


# ---------------------------------------------------------------------------
# Grade × Class → Status matrix
# ---------------------------------------------------------------------------
#
#  Class │  A      B      C      D
# ───────┼──────────────────────────────
#    1   │ Green  Green  Orange  Red
#    2   │ Green  Green  Orange  Red
#    3   │ Orange Orange Orange  Red
#    4   │ Red    Red    Red     Red
#
# Green  = Can Quote and Be Awarded
# Orange = Can Quote but Not be Awarded
# Red    = New Business on Hold

STATUS_GREEN = "Can Quote and Be Awarded"
STATUS_ORANGE = "Can Quote but Not be Awarded"
STATUS_RED = "New Business on Hold"

PANEL_ADD = "panel_add"
PANEL_ADD_EXEC = "panel_add_exec_committee"
PANEL_REJECT = "panel_reject"

STATUS_TO_PANEL = {
    STATUS_GREEN: PANEL_ADD,
    STATUS_ORANGE: PANEL_ADD_EXEC,
    STATUS_RED: PANEL_REJECT,
}


def grade_class_to_status(grade: str, class_value: int) -> str:
    """Derive supplier status from operational grade + class."""
    g = grade.upper().strip()
    c = int(class_value)
    if g == "D" or c == 4:
        return STATUS_RED
    if g in ("A", "B") and c <= 2:
        return STATUS_GREEN
    return STATUS_ORANGE  # A3, B3, C1, C2, C3


def compose_final_grade(grade: str, class_value: int) -> str:
    return f"{grade.upper()}{class_value}"


# ---------------------------------------------------------------------------
# Evaluation scheduling
# ---------------------------------------------------------------------------

FREQUENCY_DAYS: Dict[str, int] = {
    "Quarterly": 91,
    "Semi-Annual": 182,
    "Annual": 365,
}

# Month counts per frequency — used for month-exact next_evaluation_date
FREQUENCY_MONTHS: Dict[str, int] = {
    "Quarterly": 3,
    "Semi-Annual": 6,
    "Annual": 12,
}

# strategic = 3 months, global = 6 months, local = annually
SCOPE_DEFAULT_FREQUENCY: Dict[str, str] = {
    "strategic": "Quarterly",
    "global": "Semi-Annual",
    "local": "Annual",
}

DUE_SOON_DAYS = 30  # flag as "due soon" within this window


def _add_months(d: date, months: int) -> date:
    """Return d + months calendar months, clamping to the last day of the target month."""
    target_month = d.month - 1 + months
    year = d.year + target_month // 12
    month = target_month % 12 + 1
    day = min(d.day, calendar.monthrange(year, month)[1])
    return date(year, month, day)


def compute_next_evaluation_date(eval_date: date, frequency: str) -> date:
    months = FREQUENCY_MONTHS.get(frequency)
    if months:
        return _add_months(eval_date, months)
    return eval_date + timedelta(days=FREQUENCY_DAYS.get(frequency, 365))


def _has_eval_evidence(relation: SupplierSiteRelation) -> bool:
    """True when the relation carries any scorecard result (grade / class / final
    grade / score) — i.e. it was evaluated even if the date wasn't recorded."""
    return bool(
        (relation.operational_grade or "").strip()
        or (relation.final_grade or "").strip()
        or relation.class_value is not None
        or relation.last_eval_score is not None
    )


def infer_frequency(relation: SupplierSiteRelation) -> str:
    """Fall back to scope-based frequency if not explicitly set."""
    if (
        relation.evaluation_frequency
        and relation.evaluation_frequency in FREQUENCY_DAYS
    ):
        return relation.evaluation_frequency
    scope = (relation.global_status or "local").lower()
    return SCOPE_DEFAULT_FREQUENCY.get(scope, "Annual")


# Performance drift detection is not yet defined — will be designed later.
# def is_performance_drift(relation: SupplierSiteRelation) -> bool: ...


# ---------------------------------------------------------------------------
# Evaluation due query
# ---------------------------------------------------------------------------


async def get_evaluations_due(db: AsyncSession) -> List[Dict[str, Any]]:
    """
    Return all active relations with their evaluation status:
      NEVER_EVALUATED, OVERDUE, DUE_SOON, PERFORMANCE_DRIFT, UP_TO_DATE
    Sorted: NEVER_EVALUATED → OVERDUE → DUE_SOON → PERFORMANCE_DRIFT → UP_TO_DATE
    """
    today = date.today()
    due_threshold = today + timedelta(days=DUE_SOON_DAYS)

    stmt = (
        select(SupplierSiteRelation, SupplierUnit, AvocarbonSite)
        .join(
            SupplierUnit,
            SupplierUnit.id_supplier_unit == SupplierSiteRelation.id_supplier_unit,
        )
        .join(AvocarbonSite, AvocarbonSite.id_site == SupplierSiteRelation.id_site)
        .where(SupplierSiteRelation.is_deleted.is_(False))
        .where(SupplierSiteRelation.inactivated_at.is_(None))
        .where(SupplierSiteRelation.validation_status == "approved")
    )
    result = await db.execute(stmt)
    rows = result.all()

    items: List[Dict[str, Any]] = []
    PRIORITY = {
        "NEVER_EVALUATED": 0,
        "OVERDUE": 1,
        "DUE_SOON": 2,
        "MISSING_DATE": 3,
        "UP_TO_DATE": 4,
    }

    for rel, unit, site in rows:
        last = rel.last_evaluation_date
        nxt = rel.next_evaluation_date
        frequency = infer_frequency(rel)

        if last is None:
            # A grade/class/score means the relation WAS evaluated — the date is
            # just missing (common in the Monday import). Only treat it as never
            # evaluated when there's no scorecard evidence at all.
            status = "MISSING_DATE" if _has_eval_evidence(rel) else "NEVER_EVALUATED"
        elif nxt and nxt < today:
            status = "OVERDUE"
        elif nxt and nxt <= due_threshold:
            status = "DUE_SOON"
        else:
            status = "UP_TO_DATE"

        days_overdue: Optional[int] = None
        days_until_due: Optional[int] = None
        if nxt:
            delta = (today - nxt).days
            if delta > 0:
                days_overdue = delta
            else:
                days_until_due = abs(delta)

        items.append(
            {
                "relation_id": rel.id_relation,
                "relation_code": rel.relation_code,
                "unit_name": unit.supplier_name,
                "unit_id": unit.id_supplier_unit,
                "plant_name": site.site_name,
                "plant_city": site.city,
                "plant_country": site.country,
                "site_id": site.id_site,
                "current_grade": rel.operational_grade,
                "current_class": rel.class_value,
                "final_grade": rel.final_grade,
                "current_status": rel.supplier_status,
                "evaluation_frequency": frequency,
                "last_evaluation_date": last.isoformat() if last else None,
                "next_evaluation_date": nxt.isoformat() if nxt else None,
                "days_overdue": days_overdue,
                "days_until_due": days_until_due,
                "eval_status": status,
                "priority": PRIORITY.get(status, 99),
            }
        )

    items.sort(key=lambda x: (x["priority"], x.get("days_overdue") or 0), reverse=False)
    return items


# ---------------------------------------------------------------------------
# Excel / CSV template generator
# ---------------------------------------------------------------------------

TEMPLATE_COLUMNS = [
    "supplier_name",
    "plant_name",
    "evaluation_date",
    "operational_grade",
    "comments",
]

GRADE_HELP = [
    "# operational_grade: A / B / C / D",
    "# supplier_name and plant_name must match exactly — use the pre-filled template",
    "# class_value is managed by the PLD evaluation and is NOT updated here",
]


def _load_openpyxl_or_raise():
    try:
        import openpyxl

        return openpyxl
    except ImportError as exc:
        raise AppException(
            "Excel template generation is unavailable because the Excel dependency is not installed on the server.",
            status_code=500,
            error_code="EXCEL_DEPENDENCY_MISSING",
        ) from exc


def generate_csv_template() -> bytes:
    buf = io.StringIO()
    writer = csv.writer(buf)
    for h in GRADE_HELP:
        writer.writerow([h])
    writer.writerow(TEMPLATE_COLUMNS)
    # Example row
    writer.writerow(
        ["ACME-CN-001", "Lyon Plant", "2026-06-08", "B", "Quarterly review"]
    )
    return buf.getvalue().encode("utf-8-sig")


def generate_xlsx_template() -> bytes:
    openpyxl = _load_openpyxl_or_raise()
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Evaluations"

    header_fill = PatternFill("solid", fgColor="062B49")
    header_font = Font(bold=True, color="FFFFFF")
    col_widths = [28, 28, 18, 20, 40]

    for col_idx, col_name in enumerate(TEMPLATE_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = col_widths[col_idx - 1]

    # Example row
    for col_idx, val in enumerate(
        ["ACME-CN-001", "Lyon Plant", "2026-06-08", "B", "Quarterly review"], start=1
    ):
        ws.cell(row=2, column=col_idx, value=val)

    # Instructions sheet
    ws2 = wb.create_sheet("Instructions")
    instructions = [
        ("Column", "Description", "Allowed values"),
        ("supplier_name", "Unit identifier — copy from pre-filled template", "Text"),
        ("plant_name", "Avocarbon plant name — copy from pre-filled template", "Text"),
        ("evaluation_date", "Date of evaluation", "YYYY-MM-DD"),
        ("operational_grade", "Operational scorecard result", "A / B / C / D"),
        ("comments", "Optional evaluation notes", "Text"),
        ("", "", ""),
        ("Note:", "class_value is NOT in this file.", ""),
        ("", "It comes from the PLD class evaluation and is kept as-is.", ""),
        ("", "Status is computed from the new grade + existing class value.", ""),
        ("", "", ""),
        ("Status matrix:", "A1/A2/B1/B2 → Green (Can Quote and Be Awarded)", ""),
        ("", "A3/B3/C1/C2/C3 → Orange (Can Quote but Not be Awarded)", ""),
        ("", "A4/B4/C4/D1/D2/D3/D4 → Red (New Business on Hold)", ""),
    ]
    for row_data in instructions:
        ws2.append(list(row_data))
    ws2.column_dimensions["A"].width = 16
    ws2.column_dimensions["B"].width = 55
    ws2.column_dimensions["C"].width = 25

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


async def generate_prefilled_template(db: AsyncSession, due_only: bool = False) -> bytes:
    """
    Generate an Excel file pre-filled with active supplier–plant relations.

    due_only=False  → all approved relations (full panel)
    due_only=True   → only OVERDUE, DUE_SOON, NEVER_EVALUATED relations,
                      with extra context columns (status, last/next date, days overdue)
    """
    stmt = (
        select(SupplierSiteRelation, SupplierUnit, AvocarbonSite)
        .join(
            SupplierUnit,
            SupplierUnit.id_supplier_unit == SupplierSiteRelation.id_supplier_unit,
        )
        .join(AvocarbonSite, AvocarbonSite.id_site == SupplierSiteRelation.id_site)
        .where(SupplierSiteRelation.is_deleted.is_(False))
        .where(SupplierSiteRelation.inactivated_at.is_(None))
        .where(SupplierSiteRelation.validation_status == "approved")
        .order_by(AvocarbonSite.site_name, SupplierUnit.supplier_name)
    )
    result = await db.execute(stmt)
    all_rows = result.all()

    today = date.today()
    due_threshold = today + timedelta(days=DUE_SOON_DAYS)

    def _eval_status(rel: SupplierSiteRelation) -> str:
        last = rel.last_evaluation_date
        nxt = rel.next_evaluation_date
        if last is None:
            return "MISSING_DATE" if _has_eval_evidence(rel) else "NEVER_EVALUATED"
        if nxt and nxt < today:
            return "OVERDUE"
        if nxt and nxt <= due_threshold:
            return "DUE_SOON"
        return "UP_TO_DATE"

    if due_only:
        rows = [
            (rel, unit, site)
            for rel, unit, site in all_rows
            if _eval_status(rel)
            in ("OVERDUE", "DUE_SOON", "NEVER_EVALUATED", "MISSING_DATE")
        ]
        # Sort: OVERDUE first, then DUE_SOON, NEVER_EVALUATED, MISSING_DATE
        _priority = {
            "OVERDUE": 0,
            "DUE_SOON": 1,
            "NEVER_EVALUATED": 2,
            "MISSING_DATE": 3,
        }
        rows.sort(key=lambda x: _priority.get(_eval_status(x[0]), 99))
    else:
        rows = list(all_rows)

    openpyxl = _load_openpyxl_or_raise()
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Evaluations"

    header_fill = PatternFill("solid", fgColor="062B49")
    header_font = Font(bold=True, color="FFFFFF")
    locked_fill = PatternFill("solid", fgColor="F1F5F9")
    hint_font = Font(italic=True, color="94A3B8")

    # Due-only template has extra context columns (read-only, for reference)
    STATUS_LABEL = {
        "OVERDUE": "⚠ Overdue",
        "DUE_SOON": "⏳ Due soon",
        "NEVER_EVALUATED": "— Never evaluated",
        "MISSING_DATE": "✎ Missing eval date",
    }
    STATUS_COLOR = {
        "OVERDUE": "FEE2E2",       # red-100
        "DUE_SOON": "FEF3C7",      # amber-100
        "NEVER_EVALUATED": "F1F5F9",  # slate-100
        "MISSING_DATE": "CFFAFE",  # cyan-100
    }

    if due_only:
        columns = [
            ("supplier_name",       30, "",              True),
            ("plant_name",          28, "",              True),
            ("evaluation_date",     18, "YYYY-MM-DD",    False),
            ("operational_grade",   18, "A / B / C / D", False),
            ("comments",            40, "Optional notes", False),
            # context columns (locked, reference only)
            ("urgency",             18, "",              True),
            ("frequency",           16, "",              True),
            ("last_evaluation",     20, "",              True),
            ("next_due",            20, "",              True),
            ("days_overdue",        14, "",              True),
        ]
    else:
        columns = [
            ("supplier_name",   30, "",              True),
            ("plant_name",      28, "",              True),
            ("evaluation_date", 18, "YYYY-MM-DD",    False),
            ("operational_grade", 20, "A / B / C / D", False),
            ("comments",        45, "Optional notes", False),
        ]

    for col_idx, (col_name, width, hint, _locked) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    for row_idx, (rel, unit, site) in enumerate(rows, start=2):
        status = _eval_status(rel) if due_only else None

        def _write(col: int, value, locked: bool = False, fill_override=None):
            cell = ws.cell(row=row_idx, column=col, value=value)
            if locked:
                cell.fill = fill_override or locked_fill
            return cell

        if due_only:
            frequency = infer_frequency(rel)
            last = rel.last_evaluation_date
            nxt = rel.next_evaluation_date
            days_overdue = (today - nxt).days if nxt and nxt < today else None

            status_fill = PatternFill("solid", fgColor=STATUS_COLOR.get(status, "F1F5F9"))

            _write(1, unit.supplier_name, locked=True)
            _write(2, site.site_name, locked=True)
            # hint cells for user-fill columns
            for col_idx, hint in [(3, "YYYY-MM-DD"), (4, "A / B / C / D"), (5, "Optional notes")]:
                c = ws.cell(row=row_idx, column=col_idx, value=hint)
                c.font = hint_font
            # context columns
            _write(6,  STATUS_LABEL.get(status, status), locked=True, fill_override=status_fill)
            _write(7,  frequency,                         locked=True)
            _write(8,  last.isoformat() if last else "—", locked=True)
            _write(9,  nxt.isoformat() if nxt else "—",   locked=True)
            _write(10, days_overdue if days_overdue else "—", locked=True)
        else:
            _write(1, unit.supplier_name, locked=True)
            _write(2, site.site_name, locked=True)
            for col_idx, hint in [(3, "YYYY-MM-DD"), (4, "A / B / C / D"), (5, "Optional notes")]:
                c = ws.cell(row=row_idx, column=col_idx, value=hint)
                c.font = hint_font

    ws.freeze_panes = "C2"

    # Summary note at the top of due-only template
    if due_only:
        ws.insert_rows(1)
        summary_cell = ws.cell(row=1, column=1,
            value=f"Due / overdue evaluations as of {today.isoformat()} — {len(rows)} supplier(s) require action. Fill columns C–E only.")
        summary_cell.font = Font(bold=True, color="062B49")
        summary_cell.fill = PatternFill("solid", fgColor="DBEAFE")
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns))
        ws.freeze_panes = "C3"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Batch ingestion
# ---------------------------------------------------------------------------


class EvaluationRow:
    __slots__ = ("supplier_name", "plant_name", "evaluation_date", "grade", "comments")

    def __init__(
        self,
        supplier_name: str,
        plant_name: str,
        evaluation_date: date,
        grade: str,
        comments: str = "",
    ) -> None:
        self.supplier_name = supplier_name.strip()
        self.plant_name = plant_name.strip()
        self.evaluation_date = evaluation_date
        self.grade = grade.upper().strip()
        self.comments = comments


def parse_rows_from_csv(content: bytes) -> tuple[List[EvaluationRow], List[str]]:
    rows: List[EvaluationRow] = []
    errors: List[str] = []
    text = content.decode("utf-8-sig", errors="replace")
    reader = csv.DictReader(io.StringIO(text))
    for i, record in enumerate(reader, start=2):
        first_val = list(record.values())[0] if record else ""
        if first_val.startswith("#"):
            continue
        try:
            rows.append(_parse_record(dict(record), i))  # type: ignore[arg-type]
        except ValueError as exc:
            errors.append(f"Row {i}: {exc}")
    return rows, errors


def parse_rows_from_xlsx(content: bytes) -> tuple[List[EvaluationRow], List[str]]:
    try:
        import openpyxl
    except ImportError:
        return [], ["openpyxl not installed — upload a CSV file instead."]

    rows: List[EvaluationRow] = []
    errors: List[str] = []
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    ws = wb.active
    headers = [
        str(cell.value or "").strip().lower()
        for cell in next(ws.iter_rows(min_row=1, max_row=1))
    ]

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        # Keep native Python date/datetime objects as-is so _parse_record can handle them cleanly
        record: Dict[str, Any] = {}
        for i, v in enumerate(row):
            if i >= len(headers):
                break
            record[headers[i]] = (
                v  # do NOT stringify here — let _parse_record normalise
            )

        first = str(list(record.values())[0] or "") if record else ""
        if first.startswith("#") or not any(
            v is not None and str(v).strip() for v in record.values()
        ):
            continue
        try:
            rows.append(_parse_record(record, row_idx))
        except ValueError as exc:
            errors.append(f"Row {row_idx}: {exc}")
    return rows, errors


def _parse_record(record: Dict[str, Any], row_num: int) -> EvaluationRow:
    sc = str(record.get("supplier_name") or "").strip()
    pn = str(record.get("plant_name") or "").strip()
    raw_date = record.get("evaluation_date")
    grade = str(record.get("operational_grade") or "").strip().upper()
    comments = str(record.get("comments") or "").strip()

    if not sc:
        raise ValueError("supplier_name is required")
    if not pn:
        raise ValueError("plant_name is required")
    if grade not in ("A", "B", "C", "D"):
        raise ValueError(f"operational_grade must be A/B/C/D, got '{grade}'")

    # Resolve evaluation_date — handles:
    #   - Python date / datetime objects (openpyxl native)
    #   - ISO strings "2026-06-19"
    #   - Datetime strings "2026-06-19 00:00:00" (Excel artefact)
    #   - Empty / None → today
    try:
        if raw_date is None or str(raw_date).strip() == "":
            eval_date = date.today()
        elif isinstance(raw_date, datetime):
            eval_date = raw_date.date()
        elif isinstance(raw_date, date):
            eval_date = raw_date
        else:
            # Strip time component if present ("2026-06-19 00:00:00" → "2026-06-19")
            date_str = str(raw_date).strip().split(" ")[0].split("T")[0]
            eval_date = date.fromisoformat(date_str)
    except (ValueError, AttributeError):
        raise ValueError(f"evaluation_date must be YYYY-MM-DD, got '{raw_date}'")

    if eval_date > date.today():
        raise ValueError(
            f"evaluation_date cannot be in the future (got '{eval_date}'). "
            "Evaluations record past events — enter the date the assessment took place."
        )

    return EvaluationRow(sc, pn, eval_date, grade, comments)


async def ingest_batch(
    db: AsyncSession,
    rows: List[EvaluationRow],
    changed_by: str = "BATCH_UPLOAD",
    dry_run: bool = False,
) -> Dict[str, Any]:
    """
    dry_run=True  → validate all rows and resolve relations but do NOT write anything.
                    Returns the same summary so the user can review before committing.
    dry_run=False → full write (default).
    """
    """
    Process a list of parsed evaluation rows, updating relations accordingly.
    Returns a summary with per-row results.
    """
    processed: List[Dict[str, Any]] = []
    skipped: List[Dict[str, Any]] = []

    # Detect duplicate (supplier_name, plant_name) pairs within the uploaded file itself
    seen_keys: set[tuple[str, str]] = set()
    deduplicated: List[EvaluationRow] = []
    for row in rows:
        key = (row.supplier_name.lower(), row.plant_name.lower())
        if key in seen_keys:
            skipped.append(
                {
                    "supplier_name": row.supplier_name,
                    "plant_name": row.plant_name,
                    "reason": "Duplicate row in file — only the first occurrence is processed",
                }
            )
        else:
            seen_keys.add(key)
            deduplicated.append(row)
    rows = deduplicated

    for row in rows:
        # Resolve relation
        from sqlalchemy import func as sqlfunc

        stmt = (
            select(SupplierSiteRelation, SupplierUnit, AvocarbonSite)
            .join(
                SupplierUnit,
                SupplierUnit.id_supplier_unit == SupplierSiteRelation.id_supplier_unit,
            )
            .join(AvocarbonSite, AvocarbonSite.id_site == SupplierSiteRelation.id_site)
            # Case-insensitive match on both sides to be forgiving of capitalisation
            .where(
                sqlfunc.lower(SupplierUnit.supplier_name) == row.supplier_name.lower()
            )
            .where(sqlfunc.lower(AvocarbonSite.site_name) == row.plant_name.lower())
            .where(SupplierSiteRelation.is_deleted.is_(False))
            .where(SupplierSiteRelation.validation_status == "approved")
        )
        result = await db.execute(stmt)
        match = result.first()

        if match is None:
            skipped.append(
                {
                    "supplier_name": row.supplier_name,
                    "plant_name": row.plant_name,
                    "reason": "No active relation found for this unit–plant combination",
                }
            )
            continue

        relation, unit, site = match

        # class_value is NOT in the upload — use whatever is already on the relation
        existing_class = relation.class_value
        frequency = infer_frequency(relation)
        next_eval = compute_next_evaluation_date(row.evaluation_date, frequency)

        # Compute new status only if a class value is already recorded
        if existing_class:
            new_status = grade_class_to_status(row.grade, existing_class)
            new_panel = STATUS_TO_PANEL[new_status]
            final_grade = compose_final_grade(row.grade, existing_class)
        else:
            # No class yet — update grade only, keep existing status/panel/final_grade
            new_status = relation.supplier_status or "Active"
            new_panel = relation.panel_decision
            final_grade = row.grade  # temporary until PLD class is recorded

        old_grade = relation.operational_grade
        old_class = relation.class_value
        old_final = relation.final_grade
        old_status = relation.supplier_status
        old_panel = relation.panel_decision

        # Create evaluation cycle
        cycle = EvaluationCycle(
            id_relation=relation.id_relation,
            cycle_type="Operational",
            frequency=frequency,
            period_start=row.evaluation_date,
            period_end=row.evaluation_date,
            due_date=row.evaluation_date,
            cycle_status="Completed",
            launched_by=changed_by,
            launched_at=datetime.now(),
            completed_at=datetime.now(),
            comments=row.comments or f"Batch evaluation upload — {row.evaluation_date}",
        )
        db.add(cycle)
        await db.flush()

        # ScoreCard
        score_card = ScoreCard(
            id_relation=relation.id_relation,
            id_cycle=cycle.id_cycle,
            scorecard_date=row.evaluation_date,
            grade=row.grade,
            comments=row.comments or None,
            entered_by=changed_by,
        )
        db.add(score_card)

        # Classification
        classification = Classification(
            id_relation=relation.id_relation,
            id_cycle=cycle.id_cycle,
            classification_date=row.evaluation_date,
            class_value=existing_class,
            operational_grade=row.grade,
            final_grade=final_grade,
            panel_decision=new_panel,
            comments=row.comments or None,
            entered_by=changed_by,
        )
        db.add(classification)

        # Status history (only if something changed)
        status_changed = (
            old_grade != row.grade
            or old_class != existing_class
            or old_final != final_grade
            or old_panel != new_panel
        )
        if status_changed:
            history = SupplierStatusHistory(
                id_relation=relation.id_relation,
                old_status=old_status,
                new_status=new_status,
                old_class=old_class,
                new_class=existing_class,
                old_grade=old_grade,
                new_grade=row.grade,
                old_final_grade=old_final,
                new_final_grade=final_grade,
                old_panel_decision=old_panel,
                new_panel_decision=new_panel,
                change_reason=f"Batch evaluation upload — {row.evaluation_date}",
                changed_by=changed_by,
            )
            db.add(history)

        # Preserve a manually applied status override — if the relation's current
        # supplier_status was deliberately set to something different from the computed
        # value, keep the override rather than silently reverting it.
        has_active_override = (
            relation.supplier_status not in (None, "")
            and new_status not in (None, "")
            and relation.supplier_status != new_status
        )
        effective_status = relation.supplier_status if has_active_override else new_status

        # Backfill history with the status actually applied, not the computed value,
        # so the audit trail correctly reflects what happened.
        if status_changed and history:
            history.new_status = effective_status

        # Update relation
        relation.operational_grade = row.grade
        relation.class_value = existing_class
        relation.final_grade = final_grade
        relation.supplier_status = effective_status
        relation.panel_decision = new_panel
        relation.last_evaluation_date = row.evaluation_date
        relation.next_evaluation_date = next_eval
        relation.evaluation_frequency = frequency
        relation.last_status_change = datetime.now()
        if row.comments:
            relation.evaluation_comments = row.comments

        # Auto-create a development plan when grade is C or D,
        # unless one is already open/required for this relation.
        dev_plan_created = False
        if row.grade in ("C", "D") and not dry_run:
            existing_plan_stmt = select(SupplierDevelopmentPlan).where(
                SupplierDevelopmentPlan.id_relation == relation.id_relation,
                SupplierDevelopmentPlan.plan_status.in_(
                    ["Must be send", "Request sent", "Received", "Under Review"]
                ),
                SupplierDevelopmentPlan.is_deleted.is_(False),
            )
            existing_plan = (await db.execute(existing_plan_stmt)).scalar_one_or_none()

            if not existing_plan:
                # Grade D → exit D within 3 months; Grade C → exit C within 6 months
                due_months = 3 if row.grade == "D" else 6
                plan_due = row.evaluation_date + timedelta(days=due_months * 30)

                if row.grade == "D":
                    title = "Improvement Plan — Grade D (exit within 3 months)"
                    internal_note = (
                        "Supplier must submit an improvement plan targeting Grade B "
                        "within 6 months and exit Grade D within 3 months."
                    )
                else:
                    title = "Improvement Plan — Grade C (exit within 6 months)"
                    internal_note = (
                        "Supplier must submit an improvement plan immediately "
                        "to exit Grade C within 6 months."
                    )

                dev_plan = SupplierDevelopmentPlan(
                    id_relation=relation.id_relation,
                    plan_title=title,
                    plan_status="Must be send",
                    issue_date=row.evaluation_date,
                    due_date=plan_due,
                    internal_comments=internal_note,
                )
                db.add(dev_plan)
                dev_plan_created = True

                # Notify the assigned buyer — a Grade C/D supplier is a supply risk
                # that requires immediate attention; silent plan creation is not enough.
                buyer_email = relation.buyer_owner
                if buyer_email:
                    supplier_display = unit.supplier_name if unit else f"Relation #{relation.id_relation}"
                    site_display = site.site_name if site else f"Site #{relation.id_site}"
                    grade_label = "D (Exit within 3 months)" if row.grade == "D" else "C (Exit within 6 months)"
                    try:
                        await send_email(
                            subject=f"[Action Required] Grade {row.grade} Supplier — {supplier_display} · {site_display}",
                            recipients=[buyer_email],
                            body_html=f"""
<div style="font-family:Arial,sans-serif;max-width:600px;margin:0 auto">
  <div style="background:#7f1d1d;padding:20px 28px;border-radius:8px 8px 0 0">
    <h1 style="color:#fff;margin:0;font-size:18px">Supplier Risk Alert — Grade {row.grade}</h1>
    <p style="color:#fca5a5;margin:4px 0 0;font-size:13px">Avocarbon · Supplier Management</p>
  </div>
  <div style="background:#f8fafc;padding:24px 28px;border:1px solid #e2e8f0;border-top:none">
    <p style="margin:0 0 16px;font-size:14px;color:#1e293b">
      A development plan has been automatically created for a Grade {row.grade} supplier
      requiring your immediate attention.
    </p>
    <table style="width:100%;border-collapse:collapse;margin-bottom:20px;font-size:13px">
      <tr style="background:#fee2e2"><td style="padding:8px 12px;font-weight:bold;width:40%">Supplier</td><td style="padding:8px 12px">{supplier_display}</td></tr>
      <tr><td style="padding:8px 12px;font-weight:bold">Plant</td><td style="padding:8px 12px">{site_display}</td></tr>
      <tr style="background:#fee2e2"><td style="padding:8px 12px;font-weight:bold">Grade</td><td style="padding:8px 12px;color:#991b1b;font-weight:bold">{grade_label}</td></tr>
      <tr><td style="padding:8px 12px;font-weight:bold">Evaluation date</td><td style="padding:8px 12px">{row.evaluation_date.isoformat()}</td></tr>
      <tr style="background:#fee2e2"><td style="padding:8px 12px;font-weight:bold">Plan due date</td><td style="padding:8px 12px;color:#991b1b;font-weight:bold">{plan_due.isoformat()}</td></tr>
    </table>
    <p style="margin:0;font-size:13px;color:#475569">
      Please send the development plan request to the supplier and monitor their response.
      Review the supplier workspace in Avocarbon Supplier Management for full details.
    </p>
  </div>
  <div style="background:#f1f5f9;padding:12px 28px;border-radius:0 0 8px 8px;border:1px solid #e2e8f0;border-top:none">
    <p style="color:#94a3b8;font-size:11px;margin:0">Avocarbon Supplier Management Platform — automated alert</p>
  </div>
</div>""",
                        )
                    except Exception:
                        pass  # email failure must not roll back the evaluation batch

        processed.append(
            {
                "supplier_name": row.supplier_name,
                "plant_name": row.plant_name,
                "evaluation_date": row.evaluation_date.isoformat(),
                "grade": row.grade,
                "class_value": existing_class,
                "final_grade": final_grade,
                "new_status": new_status,
                "status_changed": status_changed,
                "previous_grade": old_grade,
                "previous_class": old_class,
                "next_evaluation_date": next_eval.isoformat(),
                "evaluation_frequency": frequency,
                "dev_plan_created": dev_plan_created,
            }
        )

    if dry_run:
        await db.rollback()
    else:
        await db.commit()

    return {
        "total_rows": len(rows),
        "processed": len(processed),
        "skipped": len(skipped),
        "dry_run": dry_run,
        "processed_rows": processed,
        "skipped_rows": skipped,
    }
